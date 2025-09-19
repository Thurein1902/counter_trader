import os
import time
import glob
from openpyxl import load_workbook
from datetime import datetime

# === CONFIGURATION ===
LOG_FOLDER = r"C:\Users\MT4ver2-e18-AZIzrF0D\AppData\Roaming\MetaQuotes\Terminal\7E59B46FD773C6FE7B889FC92951284D\MQL5\Files"
EXCEL_FILE = r"G:\共有ドライブ\Trading_Signal\Hourly_Signal_28pair.xlsx"
SHEET_NAME = "parameters"
COLUMN = "A"

# === FUNCTION: Read file with encoding fallback ===
def read_file_with_fallback_encoding(path):
    encodings = ['utf-8', 'utf-16', 'cp932', 'shift_jis', 'iso-8859-1']
    for enc in encodings:
        try:
            with open(path, 'r', encoding=enc) as f:
                return f.read().strip()
        except UnicodeDecodeError:
            continue
    raise ValueError("Unable to decode file with common encodings.")

# === FUNCTION: Get latest .log file ===
def get_latest_log_file():
    log_files = glob.glob(os.path.join(LOG_FOLDER, "*.log"))
    if not log_files:
        return None
    return max(log_files, key=os.path.getmtime)

# === FUNCTION: Clear and write to Excel ===
def append_log_to_excel(log_path):
    content = read_file_with_fallback_encoding(log_path)
    lines = content.splitlines()
    
    wb = load_workbook(EXCEL_FILE)
    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ Sheet '{SHEET_NAME}' not found in Excel file!")
        return
    ws = wb[SHEET_NAME]

    # Clear all data in column A
    max_row = ws.max_row
    for row in range(1, max_row + 1):
        ws[f"{COLUMN}{row}"] = None

    # Write each line to column A
    row = 1
    for line in lines:
        ws[f"{COLUMN}{row}"] = line
        row += 1

    wb.save(EXCEL_FILE)
    print(f"✅ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} — {len(lines)} lines saved")

# === MAIN LOOP: Run at 04 minute every hour ===
def run_at_minute(minute=4):
    print(f"🕒 Waiting for minute {minute:02d} of each hour to run...\n")
    last_processed_hour = None

    while True:
        now = datetime.now()
        if now.minute == minute and now.hour != last_processed_hour:
            print(f"🔄 Processing log at {now.strftime('%Y-%m-%d %H:%M')}...")
            latest_file = get_latest_log_file()
            if latest_file:
                append_log_to_excel(latest_file)
            else:
                print("⚠️ No log file found.")
            last_processed_hour = now.hour
            time.sleep(60)  # wait 1 minute to avoid double-run in same minute

        time.sleep(1)

# === ENTRY POINT ===
if __name__ == "__main__":
    run_at_minute(minute=4)
